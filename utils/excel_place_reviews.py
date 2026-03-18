from typing import List
from datetime import date
import re

from openpyxl import load_workbook

from schemas.place_insights import ReviewRowIn


KOREAN_WEEKDAY_PATTERN = re.compile(r"\.(월|화|수|목|금|토|일)$")



def parse_review_date(raw: str, default_year: int) -> date:
    s = (raw or "").strip()
    if not s:
        return date(default_year, 1, 1)

    # 1) 요일 제거
    s = KOREAN_WEEKDAY_PATTERN.sub("", s).strip()

    # 2) YYYY.MM.DD / YYYY-MM-DD / YYYY/MM/DD
    m = re.search(r"(\d{4})[.\-\/](\d{1,2})[.\-\/](\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)

    # 3) M.D / M/D
    m = re.search(r"^(\d{1,2})[.\-\/](\d{1,2})$", s)
    if m:
        mo, d = map(int, m.groups())
        return date(default_year, mo, d)

    # 4) 2월 11일
    m = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", s)
    if m:
        mo, d = map(int, m.groups())
        return date(default_year, mo, d)

    return date(default_year, 1, 1)
    
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

# 실패 코드 표준 (최소 세트)
ERR_MISSING_FIELD = "MISSING_FIELD"
ERR_INVALID_TYPE = "INVALID_TYPE"
ERR_INVALID_FORMAT = "INVALID_FORMAT"
ERR_OUT_OF_RANGE = "OUT_OF_RANGE"
ERR_TOO_LONG = "TOO_LONG"
ERR_DUPLICATE = "DUPLICATE"
ERR_PARSE_ERROR = "PARSE_ERROR"


def _norm_header(s: Any) -> str:
    return str(s or "").strip().lower()


def _failure(row: int, code: str, message: str, field: Optional[str] = None, value: Optional[Any] = None) -> Dict[str, Any]:
    return {
        "row": int(row),
        "code": code,
        "field": field,
        "message": message,
        "value": None if value is None else str(value),
    }


def _parse_xlsx_via_zip(file_bytes: bytes) -> Tuple[Optional[List[tuple]], Optional[str]]:
    """xlsx를 ZIP으로 열어 XML 직접 파싱 (스타일 깨진 파일 대응)"""
    import zipfile
    from xml.etree import ElementTree as ET

    NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    try:
        zf = zipfile.ZipFile(BytesIO(file_bytes))
    except Exception:
        return None, None

    # sharedStrings 로드
    shared = []
    if "xl/sharedStrings.xml" in zf.namelist():
        try:
            tree = ET.parse(zf.open("xl/sharedStrings.xml"))
            for si in tree.findall(f"{NS}si"):
                texts = [t.text or "" for t in si.iter(f"{NS}t")]
                shared.append("".join(texts))
        except Exception:
            pass

    # 첫 번째 시트 찾기
    sheet_file = None
    sheet_title = "Sheet1"
    for name in zf.namelist():
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            sheet_file = name
            break
    if not sheet_file:
        return None, None

    try:
        tree = ET.parse(zf.open(sheet_file))
    except Exception:
        return None, None

    rows_data = []
    for row_el in tree.iter(f"{NS}row"):
        cells = []
        for c in row_el.findall(f"{NS}c"):
            t = c.get("t", "")  # s=shared string, inlineStr, n=number, etc
            val = None
            if t == "inlineStr":
                is_el = c.find(f"{NS}is")
                if is_el is not None:
                    texts = [tt.text or "" for tt in is_el.iter(f"{NS}t")]
                    val = "".join(texts)
            elif t == "s":
                v_el = c.find(f"{NS}v")
                if v_el is not None and v_el.text is not None:
                    idx = int(v_el.text)
                    val = shared[idx] if idx < len(shared) else v_el.text
            else:
                v_el = c.find(f"{NS}v")
                if v_el is not None and v_el.text is not None:
                    try:
                        val = float(v_el.text)
                        if val == int(val):
                            val = int(val)
                    except ValueError:
                        val = v_el.text
            cells.append(val)
        if cells:
            rows_data.append(tuple(cells))

    return (rows_data if rows_data else None), sheet_title


def load_reviews_from_excel(file_bytes: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
    """
    Excel(xlsx) bytes -> parse rows -> return meta/items/failures/stats

    반환 표준:
    - meta: total/ok/fail/source
    - items: 성공 row (도메인 객체 dict)
    - failures: 실패 row 리포트 dict
    - stats.codes: 실패 코드 카운트
    """
    result: Dict[str, Any] = {
        "meta": {
            "total": 0,
            "ok": 0,
            "fail": 0,
            "source": {"filename": filename, "sheet": None},
        },
        "items": [],
        "failures": [],
        "stats": {"codes": {}},
    }

    def bump(code: str):
        codes = result["stats"]["codes"]
        codes[code] = int(codes.get(code, 0)) + 1

    # 1) 워크북 로드 — openpyxl 우선, 실패 시 ZIP/XML 직접 파싱
    all_rows = None
    sheet_title = None

    # 1-a) openpyxl 시도 (read_only 우선)
    for read_only in (True, False):
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=read_only)
            ws = wb.active
            sheet_title = ws.title
            all_rows = list(ws.iter_rows(values_only=True))
            try:
                wb.close()
            except Exception:
                pass
            break
        except Exception:
            continue

    # 1-b) openpyxl 실패 시 ZIP/XML 직접 파싱 (스타일 깨진 파일 대응)
    if all_rows is None:
        all_rows, sheet_title = _parse_xlsx_via_zip(file_bytes)

    if all_rows is None:
        f = _failure(row=0, code=ERR_PARSE_ERROR, message="엑셀 파일을 열 수 없습니다. 파일이 손상되었을 수 있습니다.")
        result["failures"].append(f)
        bump(ERR_PARSE_ERROR)
        result["meta"]["fail"] = 1
        return result

    result["meta"]["source"]["sheet"] = sheet_title
    if not all_rows:
        f = _failure(row=0, code=ERR_PARSE_ERROR, message="빈 엑셀 파일입니다.")
        result["failures"].append(f)
        bump(ERR_PARSE_ERROR)
        return result

    # 3) 헤더 읽기 (첫 줄)
    headers = [_norm_header(h) for h in all_rows[0]]

    # 4) 컬럼 매핑
    def find_col(*candidates: str) -> Optional[int]:
        for cand in candidates:
            cand = cand.strip().lower()
            for idx, h in enumerate(headers):
                if h == cand or cand in h or h in cand:
                    return idx
        return None

    col_content = find_col("content", "review", "리뷰", "리뷰내용", "내용", "후기", "텍스트", "text", "comment", "본문")
    col_rating = find_col("rating", "별점", "평점", "점수", "score", "star")
    col_date = find_col("date", "reviewed_at", "작성일", "리뷰일", "날짜", "일자", "일시")
    col_platform = find_col("platform", "채널", "플랫폼", "출처", "source")
    col_external_id = find_col("external_id", "review_id", "리뷰id", "리뷰아이디", "id")

    # content 컬럼을 못 찾으면 첫 번째 컬럼을 content로 사용
    if col_content is None:
        col_content = 0

    # 5) 본문 row 순회 (2번째 줄부터)
    total = ok = fail = 0
    seen_ids = set()

    for r_idx_0, row_values in enumerate(all_rows[1:], start=2):
        # 빈 줄 스킵
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        total += 1

        try:
            def cell_val(col_idx):
                if col_idx is not None and col_idx < len(row_values):
                    return row_values[col_idx]
                return None

            raw_content = cell_val(col_content)
            content = (str(raw_content).strip() if raw_content is not None else "")
            if not content:
                fail += 1
                f = _failure(row=r_idx_0, code=ERR_MISSING_FIELD, field="content", message="content is required", value=raw_content)
                result["failures"].append(f)
                bump(ERR_MISSING_FIELD)
                continue

            # rating (선택)
            rating_val = None
            if col_rating is not None:
                raw_rating = cell_val(col_rating)
                if raw_rating is not None and str(raw_rating).strip() != "":
                    try:
                        rating_val = float(raw_rating)
                        if rating_val < 1 or rating_val > 5:
                            fail += 1
                            f = _failure(row=r_idx_0, code=ERR_OUT_OF_RANGE, field="rating", message="rating must be 1~5", value=raw_rating)
                            result["failures"].append(f)
                            bump(ERR_OUT_OF_RANGE)
                            continue
                    except Exception:
                        fail += 1
                        f = _failure(row=r_idx_0, code=ERR_INVALID_TYPE, field="rating", message="rating must be number", value=raw_rating)
                        result["failures"].append(f)
                        bump(ERR_INVALID_TYPE)
                        continue

            # date (선택)
            reviewed_at = None
            if col_date is not None:
                raw_date = cell_val(col_date)
                if raw_date is not None and str(raw_date).strip() != "":
                    if isinstance(raw_date, date):
                        reviewed_at = raw_date.isoformat()
                    else:
                        reviewed_at = parse_review_date(str(raw_date), default_year=date.today().year).isoformat()

            platform = None
            if col_platform is not None:
                raw_platform = cell_val(col_platform)
                platform = None if raw_platform is None else str(raw_platform).strip() or None

            external_id = None
            if col_external_id is not None:
                raw_eid = cell_val(col_external_id)
                external_id = None if raw_eid is None else str(raw_eid).strip() or None
                if external_id:
                    if external_id in seen_ids:
                        fail += 1
                        f = _failure(row=r_idx_0, code=ERR_DUPLICATE, field="external_id", message="duplicate external_id", value=external_id)
                        result["failures"].append(f)
                        bump(ERR_DUPLICATE)
                        continue
                    seen_ids.add(external_id)

            item = {
                "row": r_idx_0,
                "content": content,
                "rating": rating_val,
                "reviewed_at": reviewed_at,
                "platform": platform,
                "external_id": external_id,
                "extra": {},
            }

            result["items"].append(item)
            ok += 1

        except Exception as e:
            fail += 1
            f = _failure(row=r_idx_0, code=ERR_PARSE_ERROR, message=f"row parse failed: {e}")
            result["failures"].append(f)
            bump(ERR_PARSE_ERROR)

    result["meta"]["total"] = total
    result["meta"]["ok"] = ok
    result["meta"]["fail"] = fail

    return result
   