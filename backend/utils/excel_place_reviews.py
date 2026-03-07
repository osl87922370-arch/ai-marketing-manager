from typing import List
from datetime import date
import re

from openpyxl import load_workbook

from ..schemas.place_insights import ReviewRowIn


KOREAN_WEEKDAY_PATTERN = re.compile(r"\.(월|화|수|목|금|토|일)$")



def parse_review_date(raw: str, default_year: int) -> date:
    s = (raw or "").strip()
    if not s:
        return date(default_year, 1, 1)

    # 1) 요일 제거
    s = KOREAN_WEEKDAY_PATTERN.sub("", s).strip()

    # 2) YYYY.MM.DD / YYYY-MM-DD / YYYY/MM/DD
    m = re.search(r"(\d{4})[.\-\/](\d{1,2})[.\-\/](\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)

    # 3) M.D / M/D
    m = re.search(r"^(\d{1,2})[.\-\/](\d{1,2})$", s)
    if m:
        mo, d = map(int, m.groups())
        return date(default_year, mo, d)

    # 4) 2월 11일
    m = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", s)
    if m:
        mo, d = map(int, m.groups())
        return date(default_year, mo, d)

    return date(default_year, 1, 1)
    
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

# 실패 코드 표준 (최소 세트)
ERR_MISSING_FIELD = "MISSING_FIELD"
ERR_INVALID_TYPE = "INVALID_TYPE"
ERR_INVALID_FORMAT = "INVALID_FORMAT"
ERR_OUT_OF_RANGE = "OUT_OF_RANGE"
ERR_TOO_LONG = "TOO_LONG"
ERR_DUPLICATE = "DUPLICATE"
ERR_PARSE_ERROR = "PARSE_ERROR"


def _norm_header(s: Any) -> str:
    return str(s or "").strip().lower()


def _failure(row: int, code: str, message: str, field: Optional[str] = None, value: Optional[Any] = None) -> Dict[str, Any]:
    return {
        "row": int(row),
        "code": code,
        "field": field,
        "message": message,
        "value": None if value is None else str(value),
    }


def load_reviews_from_excel(file_bytes: bytes, filename: Optional[str] = None) -> Dict[str, Any]:
    """
    Excel(xlsx) bytes -> parse rows -> return meta/items/failures/stats

    반환 표준:
    - meta: total/ok/fail/source
    - items: 성공 row (도메인 객체 dict)
    - failures: 실패 row 리포트 dict
    - stats.codes: 실패 코드 카운트
    """
    result: Dict[str, Any] = {
        "meta": {
            "total": 0,
            "ok": 0,
            "fail": 0,
            "source": {"filename": filename, "sheet": None},
        },
        "items": [],
        "failures": [],
        "stats": {"codes": {}},
    }

    def bump(code: str):
        codes = result["stats"]["codes"]
        codes[code] = int(codes.get(code, 0)) + 1

    # 1) 워크북 로드 (bytes)
    try:
        wb = load_workbook(
    BytesIO(file_bytes),
    data_only=True,
    read_only=True
)
    except Exception as e:
        f = _failure(row=0, code=ERR_PARSE_ERROR, message=f"workbook load failed: {e}")
        result["failures"].append(f)
        bump(ERR_PARSE_ERROR)
        result["meta"]["fail"] = 1
        return result

    ws = wb.active
    result["meta"]["source"]["sheet"] = ws.title

    # 2) 헤더 읽기 (첫 줄)
    header_row_idx = 1
    
    headers_raw = []
    for col in range(1, ws.max_column + 1):
        headers_raw.append(ws.cell(row=1, column=col).value or "")
    headers = [_norm_header(h) for h in headers_raw]

    # 3) 컬럼 매핑: 최소로 content만 필수로 잡고, 나머지는 있으면 읽는다
    #    (엑셀 템플릿이 정해져 있으면 여기서 매핑을 더 타이트하게 만들면 됨)
    def find_col(*candidates: str) -> Optional[int]:
        for cand in candidates:
            cand = cand.strip().lower()
            if cand in headers:
                return headers.index(cand)  # 0-based
        return None

    col_content = find_col("content", "review", "리뷰", "리뷰내용", "내용")
    col_rating = find_col("rating", "별점", "평점")
    col_date = find_col("date", "reviewed_at", "작성일", "리뷰일", "날짜")
    col_platform = find_col("platform", "채널", "플랫폼")
    col_external_id = find_col("external_id", "review_id", "리뷰id", "리뷰아이디", "id")

    if col_content is None:
        f = _failure(row=1, code=ERR_MISSING_FIELD, field="content", message="header missing: content (or 리뷰내용/내용)")
        result["failures"].append(f)
        bump(ERR_MISSING_FIELD)
        result["meta"]["fail"] = 1
        return result

    # 4) 본문 row 순회 (2번째 줄부터)
    #    row 번호는 엑셀 기준(1-index) 그대로 반환
    total = ok = fail = 0
    seen_ids = set()

    for r_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = []
        for col in range(1, ws.max_column + 1):
            row_cells.append(ws.cell(row=r_idx, column=col))
        # 빈 줄 스킵(모든 셀이 비어있으면)
        if all((c.value is None or str(c.value).strip() == "") for c in row_cells):
            continue

        total += 1

        try:
            raw_content = row_cells[col_content].value
            content = (str(raw_content).strip() if raw_content is not None else "")
            if not content:
                fail += 1
                f = _failure(row=r_idx, code=ERR_MISSING_FIELD, field="content", message="content is required", value=raw_content)
                result["failures"].append(f)
                bump(ERR_MISSING_FIELD)
                continue

            # rating (선택)
            rating_val = None
            if col_rating is not None:
                raw_rating = row_cells[col_rating].value
                if raw_rating is not None and str(raw_rating).strip() != "":
                    try:
                        rating_val = float(raw_rating)
                        # 예: 1~5 범위 가드(원하면 조정)
                        if rating_val < 1 or rating_val > 5:
                            fail += 1
                            f = _failure(row=r_idx, code=ERR_OUT_OF_RANGE, field="rating", message="rating must be 1~5", value=raw_rating)
                            result["failures"].append(f)
                            bump(ERR_OUT_OF_RANGE)
                            continue
                    except Exception:
                        fail += 1
                        f = _failure(row=r_idx, code=ERR_INVALID_TYPE, field="rating", message="rating must be number", value=raw_rating)
                        result["failures"].append(f)
                        bump(ERR_INVALID_TYPE)
                        continue

            # date (선택) - 너 파일에 있는 parse_review_date() 활용
            reviewed_at = None
            if col_date is not None:
                raw_date = row_cells[col_date].value
                if raw_date is not None and str(raw_date).strip() != "":
                    if isinstance(raw_date, date):
                        reviewed_at = raw_date.isoformat()
                    else:
                        # default_year는 올해로 (원하면 변경)
                        reviewed_at = parse_review_date(str(raw_date), default_year=date.today().year).isoformat()

            platform = None
            if col_platform is not None:
                raw_platform = row_cells[col_platform].value
                platform = None if raw_platform is None else str(raw_platform).strip() or None

            external_id = None
            if col_external_id is not None:
                raw_eid = row_cells[col_external_id].value
                external_id = None if raw_eid is None else str(raw_eid).strip() or None
                if external_id:
                    if external_id in seen_ids:
                        fail += 1
                        f = _failure(row=r_idx, code=ERR_DUPLICATE, field="external_id", message="duplicate external_id", value=external_id)
                        result["failures"].append(f)
                        bump(ERR_DUPLICATE)
                        continue
                    seen_ids.add(external_id)

            item = {
                "row": r_idx,
                "content": content,
                "rating": rating_val,
                "reviewed_at": reviewed_at,
                "platform": platform,
                "external_id": external_id,
                "extra": {},
            }

            result["items"].append(item)
            ok += 1

        except Exception as e:
            fail += 1
            f = _failure(row=r_idx, code=ERR_PARSE_ERROR, message=f"row parse failed: {e}")
            result["failures"].append(f)
            bump(ERR_PARSE_ERROR)

    result["meta"]["total"] = total
    result["meta"]["ok"] = ok
    result["meta"]["fail"] = fail

    return result
   